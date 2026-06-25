import json
import os
import re
from flask import Flask, render_template, request, session, redirect, url_for, jsonify

# ============================================================================
# SCRAPER CONFIGURATION
# ============================================================================
# Default ke Apify dulu, user bisa switch ke Smartproxy dengan mengubah import
try:
    # Option A: Apify (Existing)
    from apify_scraper import ApifyInstagramScraper
    
    # Option B: Smartproxy (New) - Uncomment to use if you have credentials
    # from smartproxy_scraper import SmartproxyInstagramScraper as ApifyInstagramScraper
    
    SCRAPER_AVAILABLE = True
except ImportError:
    SCRAPER_AVAILABLE = False

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'dev-secret-change-me')

import database
database.init_db()

# Subpath configuration for VPS deployment
# When deployed at /kolproject, set this. For local dev, leave empty.
SUBPATH = os.environ.get('APPLICATION_ROOT', '')
if SUBPATH:
    from werkzeug.middleware.dispatcher import DispatcherMiddleware
    from werkzeug.wrappers import Response
    app.config['APPLICATION_ROOT'] = SUBPATH


def load_data():
    file_path = 'influencers.json'
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def load_regions_data():
    file_path = 'data/indonesia_cities.json'
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"provinces": []}

REGIONS_DB = load_regions_data()
AUTO_SCRAPE_MIN_LOCAL_RESULTS = int(os.environ.get('AUTO_SCRAPE_MIN_LOCAL_RESULTS', '5'))
AUTO_SCRAPE_MAX_HASHTAGS = int(os.environ.get('AUTO_SCRAPE_MAX_HASHTAGS', '2'))
AUTO_SCRAPE_LIMIT_PER_TAG = int(os.environ.get('AUTO_SCRAPE_LIMIT_PER_TAG', '20'))
AUTO_SCRAPE_LOCATION_LIMIT = int(os.environ.get('AUTO_SCRAPE_LOCATION_LIMIT', '20'))
AUTO_SCRAPE_PROFILE_LIMIT = int(os.environ.get('AUTO_SCRAPE_PROFILE_LIMIT', '5'))
AUTO_SCRAPE_REQUIRE_CITY = os.environ.get('AUTO_SCRAPE_REQUIRE_CITY', '1') == '1'


def normalize_province_name(value):
    return (
        (value or '')
        .lower()
        .strip()
        .replace("dki ", "")
        .replace("d.i. ", "")
        .replace("di ", "")
        .replace("daerah istimewa ", "")
    )


def parse_followers_count(value):
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)

    text = str(value or '').strip().upper().replace(',', '')
    if not text:
        return 0

    multiplier = 1
    if text.endswith('K'):
        multiplier = 1000
        text = text[:-1]
    elif text.endswith('M'):
        multiplier = 1000000
        text = text[:-1]

    try:
        return int(float(text) * multiplier)
    except ValueError:
        digits_only = ''.join(ch for ch in text if ch.isdigit())
        return int(digits_only) if digits_only else 0


def parse_engagement_rate(value):
    text = str(value or '0').replace('%', '').replace(',', '.').strip()
    try:
        return float(text)
    except ValueError:
        return 0.0


def influencer_matches_filters(influencer, filters, raw_query, clean_query, niche_list):
    name = influencer.get('name', '').lower()
    username = influencer.get('username', '').lower()
    bio = influencer.get('bio', '').lower()
    tags = [t.lower() for t in influencer.get('tags', [])]
    location = influencer.get('location', '').lower()
    province = (influencer.get('location_province') or '').lower().strip()
    detected = [d.lower() for d in influencer.get('detected_locations', [])]

    # === FILTER MERCHANT (Toko/Bisnis/Restoran) ===
    # Buang akun yang secara eksplisit adalah tempat usaha, bukan KOL/Influencer,
    # kecuali user benar-benar mencari kata kunci spesifik toko.
    import re
    merchant_keywords = [
        r'\bgrosir\b', r'\brepair\b', r'\bservice\b', r'\bgaransi\b', r'\bknalpot\b', r'\barloji\b',
        r'\bjam tangan\b', r'\bbengkel\b', r'\bsupplier\b', r'ready\s?stock', r'frozen\s?food', 
        r'\bcatering\b', r'jual beli', r'open order', r'close order', r'delivery order', 
        r'\bgofood\b', r'\bgrabfood\b', r'\bshopeefood\b', r'setiap hari', r'pemesanan via',
        r'menerima pesanan', r'melayani pesanan', r'\bwarung\b', r'\bkedai\b', r'\brestoran\b', 
        r'\bresto\b', r'\bbuka\s+\d{1,2}', r'\bapotek\b', r'\bklinik\b', r'\bpenginapan\b', r'\bvilla\b'
    ]
    bio_name_str = f"{name} {username} {bio}"
    for kw in merchant_keywords:
        if re.search(kw, bio_name_str):
            if not raw_query or not re.search(kw, raw_query.lower()):
                return False
    if raw_query:
        text_match = any(
            query in field
            for query in {raw_query, clean_query}
            if query
            for field in (name, username, bio)
        )
        tag_match = any(clean_query == tag or clean_query in tag for tag in tags) if clean_query else False
        if not text_match and not tag_match:
            return False

    if filters["province"]:
        if not province:
            return False
        if normalize_province_name(filters["province"]) != normalize_province_name(province):
            return False

    if filters["city"]:
        city_filter = filters["city"]
        city_match = (
            city_filter in location
            or location in city_filter
            or city_filter in province
            or province in city_filter
            or any(city_filter in item or item in city_filter for item in detected)
        )
        if not city_match:
            return False

    if filters["district"]:
        district_filter = filters["district"]
        district_match = (
            any(district_filter in item or item in district_filter for item in detected)
            or district_filter in location
            or location in district_filter
        )
        if not district_match:
            return False

    if filters["min_fol"]:
        try:
            min_followers = int(filters["min_fol"])
        except ValueError:
            min_followers = 0
        current_followers = influencer.get('followers_int') or parse_followers_count(influencer.get('followers'))
        if current_followers < min_followers:
            return False

    if filters.get("max_fol"):
        try:
            max_followers = int(filters["max_fol"])
        except ValueError:
            max_followers = float('inf')
        current_followers = influencer.get('followers_int') or parse_followers_count(influencer.get('followers'))
        if current_followers > max_followers:
            return False

    if filters["er_min"] and filters["er_min"] != "0":
        if parse_engagement_rate(influencer.get('er')) < parse_engagement_rate(filters["er_min"]):
            return False

    if niche_list:
        # User request: "gunakan hastag pencaharian dari postingan feed atau reels saja"
        caption_hashtags = [h.lower() for h in influencer.get('caption_hashtags', [])]

        niche_match = False
        for niche in niche_list:
            niche_raw_lower = niche.strip().lower().replace('#', '')
            niche_clean = niche_raw_lower.replace(' ', '')
            if not niche_clean:
                continue

            # 1. Cari di caption_hashtags (substring match bebas untuk hashtag)
            if any(niche_clean in h for h in caption_hashtags):
                niche_match = True
                break
                
            # 2. Cari langsung di Bio (harus exact word match agar tidak salah target)
            bio_lower = influencer.get('bio', '').lower()
            if re.search(r'\b' + re.escape(niche_raw_lower) + r'\b', bio_lower):
                niche_match = True
                break
            # Coba juga format tanpa spasi (misal bio nulis 'malangfoodies')
            if niche_clean != niche_raw_lower and re.search(r'\b' + re.escape(niche_clean) + r'\b', bio_lower):
                niche_match = True
                break
            
            # Jika user memasukkan spasi, misal "kuliner malang", maka pisah per kata 
            # dan pastikan SEMUA kata ada di dalam daftar hashtag
            if ' ' in niche.strip():
                words = niche.strip().lower().replace('#', '').split()
                # Cek apakah setiap kata ada di setidaknya satu hashtag
                all_words_found = True
                for w in words:
                    if not any(w in h for h in caption_hashtags):
                        all_words_found = False
                        break
            # 1. Cari di caption_hashtags
            if any(niche_clean in h for h in caption_hashtags):
                niche_match = True
                break
                
            # 2. LOGIKA CERDAS: Fallback untuk kategori Influencer / Artis / Content Creator
            if niche_clean in ['influencermalang', 'selebgrammalang', 'malangcontentcreator', 'contentcreatormalang', 'artismalang']:
                bio_lower = influencer.get('bio', '').lower()
                name_lower = (influencer.get('full_name', '') + ' ' + influencer.get('username', '')).lower()
                
                # Hindari salon / MUA / entitas bisnis masuk daftar influencer secara tidak sengaja
                is_business = any(re.search(r'\b' + m + r'\b', bio_lower + ' ' + name_lower) for m in ['mua', 'makeup', 'salon', 'eyelash', 'nail', 'hotel', 'hostel', 'villa', 'penginapan', 'univ', 'universitas', 'kampus', 'boutique', 'butik', 'toko', 'store', 'jual', 'wisata', 'cafe', 'resto', 'tour', 'travel'])
                
                if not is_business:
                    # Gunakan regex word boundary agar 'host' tidak cocok dengan 'hostel'
                    professions = ['influencer', 'selebgram', 'public figure', 'content creator', 'digital creator', 'video creator', 'host', 'mc']
                    if any(re.search(r'\b' + kw + r'\b', bio_lower) for kw in professions):
                        niche_match = True
                        break
                        
                    # Jika followers > 10000 dan tipe Creator, otomatis dianggap influencer besar
                    try:
                        followers = int(influencer.get('followers', 0) or 0)
                    except (ValueError, TypeError):
                        followers = 0
                    if followers >= 10000 and influencer.get('account_type', '').lower() == 'creator':
                        niche_match = True
                        break
                    
            # 3. LOGIKA CERDAS: Fallback untuk Standup Comedy
            if niche_clean in ['standupcomedymalang', 'komikamalang', 'komedimalang']:
                bio_lower = influencer.get('bio', '').lower()
                if any(kw in bio_lower for kw in ['standup', 'komika', 'comedian', 'komedi', 'pelawak', 'comedy', 'lucu']):
                    niche_match = True
                    break

        if not niche_match:
            return False

    if filters.get("account_type"):
        acc_type_filter = filters["account_type"].lower()
        if influencer.get("account_type", "").lower() != acc_type_filter:
            return False

    return True


def filter_influencers(influencers, filters, raw_query, clean_query, niche_list):
    return [
        influencer
        for influencer in influencers
        if influencer_matches_filters(influencer, filters, raw_query, clean_query, niche_list)
    ]


def build_scrape_terms(filters, clean_query, niche_list):
    scrape_terms = []

    if clean_query:
        scrape_terms.append(clean_query)

    if filters["province"]:
        province_clean = filters["province"].lower().replace(" ", "").replace("dki", "")
        scrape_terms.extend([province_clean, f"explore{province_clean}"])

    if filters["city"]:
        city_clean = (
            filters["city"]
            .lower()
            .replace("kota ", "")
            .replace("kabupaten ", "")
            .replace(" ", "")
        )
        scrape_terms.extend([city_clean, f"explore{city_clean}"])

    for niche in niche_list[:3]:
        niche_clean = niche.strip().replace("#", "").replace(" ", "")
        if niche_clean and len(niche_clean) > 2:
            scrape_terms.append(niche_clean)

    return sorted(set(scrape_terms))[:AUTO_SCRAPE_MAX_HASHTAGS]


def dedupe_influencers(influencers):
    unique = []
    seen = set()
    for influencer in influencers:
        username = influencer.get('username')
        if not username or username in seen:
            continue
        unique.append(influencer)
        seen.add(username)
    return unique


def maybe_auto_scrape(filtered_data, filters, scrape_terms, niche_list, raw_query, clean_query):
    has_precise_location = bool(filters["city"])
    should_scrape = (
        len(filtered_data) < AUTO_SCRAPE_MIN_LOCAL_RESULTS
        and bool(scrape_terms)
        and SCRAPER_AVAILABLE
        and (has_precise_location or not AUTO_SCRAPE_REQUIRE_CITY)
    )
    print(f"[DEBUG] filtered_data count: {len(filtered_data)}")
    print(f"[DEBUG] scrape_terms: {scrape_terms}")
    print(f"[DEBUG] SCRAPER_AVAILABLE: {SCRAPER_AVAILABLE}")

    if not should_scrape:
        print(f"[DEBUG] Scraping skipped - should_scrape={should_scrape}")
        return filtered_data

    print(f"[INFO] Data minim ({len(filtered_data)}). Auto-Scrape dengan terms: {scrape_terms}")
    try:
        scraper = ApifyInstagramScraper()
        all_new_kols = []

        if filters["city"]:
            city_name = filters["city"].replace("kota ", "").replace("kabupaten ", "").title()
            print(f"[SCRAPE] Scraping by location: {city_name}...")
            try:
                location_kols = scraper.scrape_location(
                    city_name,
                    limit=AUTO_SCRAPE_LOCATION_LIMIT,
                    max_profiles=AUTO_SCRAPE_PROFILE_LIMIT,
                )
                if location_kols:
                    all_new_kols.extend(location_kols)
                    print(f"[OK] Got {len(location_kols)} KOLs from location search")
            except Exception as loc_err:
                print(f"[WARN] Location scrape error: {loc_err}")

        if len(filtered_data) + len(all_new_kols) < AUTO_SCRAPE_MIN_LOCAL_RESULTS and scrape_terms:
            terms_to_scrape = scrape_terms[:AUTO_SCRAPE_MAX_HASHTAGS]
            print(f"[SCRAPE] Batch Scraping hashtags: #{', #'.join(terms_to_scrape)}...")
            hashtag_kols = scraper.scrape_hashtags_batch(
                terms_to_scrape,
                limit_per_tag=AUTO_SCRAPE_LIMIT_PER_TAG,
                max_profiles=AUTO_SCRAPE_PROFILE_LIMIT,
            )
            if hashtag_kols:
                all_new_kols.extend(hashtag_kols)
                print(f"[OK] Got {len(hashtag_kols)} KOLs from hashtags")

        if not all_new_kols:
            print("[WARN] No new KOLs found from scraping")
            return filtered_data

        scraper.save_results(all_new_kols)
        matched_new_kols = filter_influencers(all_new_kols, filters, raw_query, clean_query, niche_list)
        combined_results = dedupe_influencers(filtered_data + matched_new_kols)
        print(f"[OK] Auto-scrape done. Total matched: {len(combined_results)}")
        return combined_results
    except Exception as err:
        print(f"[ERROR] Auto-scrape error: {err}")
        return filtered_data


def build_light_regions(regions_db):
    light_regions = {"provinces": []}
    for prov in regions_db.get("provinces", []):
        light_regions["provinces"].append({
            "name": prov["name"],
            "cities": [{"name": city["name"]} for city in prov.get("cities", [])]
        })
    return light_regions

# ============================================================================
# ROUTES
# ============================================================================

ITEMS_PER_PAGE = 100

@app.route('/', methods=['GET', 'POST'])
def home():
    filters = {
        "query": "", "province": "", "city": "", "district": "",
        "min_fol": "", "max_fol": "", "er_min": "0", "platform": "instagram", "niche_tags": "",
        "sort_by": "", "account_type": ""
    }
    search_performed = False
    filtered_data = []
    pagination = None

    has_params = any(request.args.get(k) for k in [
        'search_keyword', 'province', 'city', 'district',
        'min_followers', 'max_followers', 'er_rate', 'niche_tags', 'sort_by', 'page',
        'account_type'
    ])

    if has_params:
        search_performed = True
        influencers_db = load_data()
        
        raw_query = request.args.get('search_keyword', '').lower().strip()
        clean_query = raw_query.replace('#', '')
        filters["query"] = raw_query
        
        filters["province"] = request.args.get('province', '')
        filters["city"] = request.args.get('city', '').lower()
        filters["district"] = request.args.get('district', '').lower()
        filters["min_fol"] = request.args.get('min_followers', '')
        filters["max_fol"] = request.args.get('max_followers', '')
        filters["er_min"] = request.args.get('er_rate', '0')
        filters["platform"] = request.args.get('platform', 'instagram')
        filters["niche_tags"] = request.args.get('niche_tags', '').lower()
        filters["sort_by"] = request.args.get('sort_by', '')
        filters["account_type"] = request.args.get('account_type', '').strip()
        niche_list = [tag.strip().replace('#', '') for tag in filters["niche_tags"].split(',') if tag.strip()]
        filtered_data = filter_influencers(influencers_db, filters, raw_query, clean_query, niche_list)
        
        if filters["sort_by"] == "followers_desc":
            filtered_data.sort(key=lambda x: x.get('followers_int') or parse_followers_count(x.get('followers')), reverse=True)
        elif filters["sort_by"] == "followers_asc":
            filtered_data.sort(key=lambda x: x.get('followers_int') or parse_followers_count(x.get('followers')))
        elif filters["sort_by"] == "er_desc":
            filtered_data.sort(key=lambda x: parse_engagement_rate(x.get('er')), reverse=True)

        total_results = len(filtered_data)
        total_pages = max(1, (total_results + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE)
        try:
            current_page = int(request.args.get('page', 1))
        except (ValueError, TypeError):
            current_page = 1
        current_page = max(1, min(current_page, total_pages))

        start_idx = (current_page - 1) * ITEMS_PER_PAGE
        end_idx = start_idx + ITEMS_PER_PAGE
        paginated_data = filtered_data[start_idx:end_idx]

        pagination = {
            "page": current_page,
            "total_pages": total_pages,
            "total_results": total_results,
            "per_page": ITEMS_PER_PAGE,
            "has_prev": current_page > 1,
            "has_next": current_page < total_pages,
        }
        filtered_data = paginated_data

    light_regions = build_light_regions(REGIONS_DB)
    if 'user_id' in session:
        saved_usernames = database.get_all_saved_usernames_by_user(session['user_id'])
    else:
        saved_usernames = {i['username'] for i in session.get('my_list', [])}
    return render_template('index.html', influencers=filtered_data, f=filters, regions=light_regions, search_performed=search_performed, pagination=pagination, saved_usernames=saved_usernames)

# ... (Sisa Routes: api/districts, influencer, add_to_list, dll TETAP SAMA) ...
@app.route('/api/districts/<city_name>')
def get_districts(city_name):
    # ... code ...
    city_name_lower = city_name.lower()
    for prov in REGIONS_DB.get("provinces", []):
        for city in prov.get("cities", []):
            if city["name"].lower() == city_name_lower:
                return {"districts": city.get("kecamatan", [])}
    return {"districts": []}

@app.route('/influencer/<username>')
def profile_detail(username):
    all_data = load_data()
    infl = next((i for i in all_data if i['username'].replace('@','') == username.replace('@','')), None)
    if infl:
        if 'user_id' in session:
            saved_usernames = database.get_all_saved_usernames_by_user(session['user_id'])
        else:
            saved_usernames = {i['username'] for i in session.get('my_list', [])}
        return render_template('profile.html', kol=infl, saved_usernames=saved_usernames)
    return "Not Found", 404

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        
        if not username or not password:
            return render_template('register.html', error='Username dan password wajib diisi.', username_val=username)
            
        if len(password) < 6:
            return render_template('register.html', error='Password minimal harus 6 karakter.', username_val=username)
            
        if password != confirm_password:
            return render_template('register.html', error='Konfirmasi password tidak cocok.', username_val=username)
            
        import database
        user_id = database.create_user(username, password)
        if user_id:
            return render_template('login.html', success='Pendaftaran berhasil! Silakan masuk dengan akun Anda.', username_val=username)
        else:
            return render_template('register.html', error='Username sudah digunakan.', username_val=username)
            
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    next_url = request.args.get('next') or request.form.get('next') or ''
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        # Check if admin login
        if username == 'admin' and password == 'admin123':
            session['is_admin'] = True
            session['username'] = 'admin'
            session['user_id'] = 9999
            return redirect(next_url or url_for('admin_dashboard'))
            
        import database
        user = database.verify_user(username, password)
        if user:
            session['user_id'] = user['id']
            session['username'] = user['username']
            # Migrate any session-based list items
            if 'my_list' in session and session['my_list']:
                lists = database.get_user_lists(user['id'])
                if lists:
                    default_list_id = lists[0]['id']
                    for kol in session['my_list']:
                        database.add_to_list(default_list_id, kol['username'])
                session.pop('my_list', None)
                
            return redirect(next_url or url_for('home'))
        else:
            return render_template('login.html', error='Username atau password salah.', next_url=next_url, username_val=username)
            
    return render_template('login.html', next_url=next_url)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('home'))

@app.route('/add_to_list/<username>')
def add_to_list(username):
    if 'user_id' not in session:
        return redirect(url_for('login', next=request.path))
        
    user_id = session['user_id']
    import database
    folders = database.get_user_lists(user_id)
    if folders:
        list_id = request.args.get('list_id')
        active_folder = None
        if list_id:
            active_folder = next((f for f in folders if str(f['id']) == str(list_id)), None)
        if not active_folder:
            active_folder = folders[0]
            
        database.add_to_list(active_folder['id'], username)
        
    return redirect(request.referrer or url_for('home'))

@app.route('/saved-lists')
def saved_lists():
    if 'user_id' not in session:
        return redirect(url_for('login', next=request.path))
        
    user_id = session['user_id']
    import database
    folders = database.get_user_lists(user_id)
    
    active_folder_id = request.args.get('list_id')
    active_folder = None
    if active_folder_id:
        active_folder = next((f for f in folders if str(f['id']) == str(active_folder_id)), None)
    if not active_folder and folders:
        active_folder = folders[0]
        
    items = []
    if active_folder:
        items = database.get_list_items(active_folder['id'])
        
    all_influencers = load_data()
    influencer_map = {i['username']: i for i in all_influencers}
    
    merged_items = []
    for item in items:
        username = item['username']
        if username in influencer_map:
            infl_detail = dict(influencer_map[username])
            infl_detail.update(item)
            merged_items.append(infl_detail)
            
    return render_template('saved_lists.html', 
                           folders=folders, 
                           active_folder=active_folder, 
                           influencers=merged_items)

@app.route('/remove_from_list/<username>')
def remove_from_list(username):
    if 'user_id' not in session:
        return redirect(url_for('login', next=request.path))
        
    user_id = session['user_id']
    import database
    
    list_id = request.args.get('list_id')
    if list_id:
        database.remove_from_list(list_id, username)
    else:
        folders = database.get_user_lists(user_id)
        for folder in folders:
            database.remove_from_list(folder['id'], username)
            
    ref = request.referrer
    if ref and 'saved-lists' in ref:
        return redirect(url_for('saved_lists', list_id=list_id))
    return redirect(ref or url_for('home'))

@app.route('/clear_list')
def clear_list():
    if 'user_id' not in session:
        return redirect(url_for('login'))
        
    list_id = request.args.get('list_id')
    if list_id:
        import database
        folders = database.get_user_lists(session['user_id'])
        if any(str(f['id']) == str(list_id) for f in folders):
            items = database.get_list_items(list_id)
            for item in items:
                database.remove_from_list(list_id, item['username'])
                
    return redirect(url_for('saved_lists', list_id=list_id))

@app.route('/create_list', methods=['POST'])
def create_list():
    if 'user_id' not in session:
        return redirect(url_for('login'))
        
    list_name = request.form.get('list_name', '').strip()
    if list_name:
        import database
        database.create_list(session['user_id'], list_name)
    return redirect(url_for('saved_lists'))

@app.route('/delete_list/<int:list_id>', methods=['POST'])
def delete_list(list_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
        
    import database
    database.delete_list(session['user_id'], list_id)
    return redirect(url_for('saved_lists'))

@app.route('/update_item_status', methods=['POST'])
def update_item_status():
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
        
    data = request.get_json() or {}
    list_id = data.get('list_id')
    username = data.get('username')
    status = data.get('status')
    notes = data.get('notes')
    
    if not list_id or not username:
        return jsonify({'status': 'error', 'message': 'Missing parameters'}), 400
        
    import database
    folders = database.get_user_lists(session['user_id'])
    if not any(str(f['id']) == str(list_id) for f in folders):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
        
    success = database.update_item_crm(list_id, username, status, notes)
    if success:
        return jsonify({'status': 'success'})
    else:
        return jsonify({'status': 'error', 'message': 'Database update failed'}), 500

@app.route('/export_list/<int:list_id>')
def export_list(list_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
        
    import database
    folders = database.get_user_lists(session['user_id'])
    folder = next((f for f in folders if f['id'] == list_id), None)
    if not folder:
        return "Access denied", 403
        
    items = database.get_list_items(list_id)
    all_influencers = load_data()
    influencer_map = {i['username']: i for i in all_influencers}
    
    import io
    import csv
    from flask import Response
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(['Username', 'Nama', 'Followers', 'ER', 'Niche', 'Email', 'No WA', 'Lokasi', 'Status', 'Catatan'])
    
    for item in items:
        username = item['username']
        if username in influencer_map:
            inf = influencer_map[username]
            writer.writerow([
                inf.get('username', ''),
                inf.get('name', ''),
                inf.get('followers', ''),
                inf.get('er', ''),
                ', '.join(inf.get('tags', [])),
                inf.get('email', ''),
                inf.get('phone', ''),
                inf.get('location', ''),
                item.get('status', 'Scouted'),
                item.get('notes', '')
            ])
            
    csv_data = output.getvalue()
    filename = f"KOL_Scout_{folder['name'].replace(' ', '_')}.csv"
    return Response(
        csv_data,
        mimetype="text/csv",
        headers={"Content-disposition": f"attachment; filename={filename}"}
    )

@app.route('/export_excel/<int:list_id>')
def export_excel(list_id):
    is_client = 'user_id' not in session
    
    import database
    with database.get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM saved_lists WHERE id = ?", (list_id,))
        folder = cursor.fetchone()
        
    if not folder:
        return "Proposal/Folder not found", 404
        
    folder = dict(folder)
    
    # If the user is logged in, verify list ownership.
    # Otherwise, it's a client proposal view access.
    if not is_client:
        folders = database.get_user_lists(session['user_id'])
        if not any(f['id'] == list_id for f in folders):
            return "Access denied", 403
            
    items = database.get_list_items(list_id)
    all_influencers = load_data()
    influencer_map = {i['username']: i for i in all_influencers}
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    import datetime
    from flask import send_file
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KOL Scout Details"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name='Segoe UI', size=16, bold=True, color='1F4E78')
    subtitle_font = Font(name='Segoe UI', size=10, italic=True, color='595959')
    header_font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Deep blue/navy
    
    data_font = Font(name='Segoe UI', size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Title Block
    ws.merge_cells('A1:S1')
    ws['A1'] = f"LAPORAN SCOUTING KOL - {folder['name'].upper()}"
    ws['A1'].font = title_font
    
    ws.merge_cells('A2:S2')
    ws['A2'] = f"Tanggal Ekspor: {datetime.datetime.now().strftime('%d %B %Y %H:%M:%S')} | Total KOL: {len(items)}"
    ws['A2'].font = subtitle_font
    
    # Blank row
    ws.append([])
    
    # Headers
    headers = [
        'Username', 'Nama', 'Followers', 'ER', 'Sponsor Posts', 'Sponsor ER',
        'Avg Views (Sponsor)', 'Avg Views (General)', 
        'Rate Reel (Rp)', 'Rate Story (Rp)', 'Rate Feed (Rp)', 
        'CPV (Rp)', 'CPE (Rp)', 'Niche', 'Email', 'No WA', 'Lokasi', 'Status CRM', 'Catatan'
    ]
    
    ws.append(headers)
    header_row_idx = 4
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row_idx, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        
    ws.row_dimensions[header_row_idx].height = 28
    
    # Data rows
    for row_idx, item in enumerate(items, 5):
        username = item['username']
        inf = influencer_map.get(username, {})
        
        # Calculate values dynamically
        followers_int = inf.get('followers_int', 0)
        
        # Parse ER
        er_val = None
        er_float = inf.get('er_float')
        if er_float is not None:
            try:
                er_val = float(er_float) / 100.0
            except ValueError:
                pass
        
        endo_posts = inf.get('endorsement_posts_count', 0)
        
        # Parse Sponsor ER
        endo_er_val = None
        endo_er_float = inf.get('endorsement_er_float')
        if endo_er_float is not None:
            try:
                endo_er_val = float(endo_er_float) / 100.0
            except ValueError:
                pass
                
        endo_avg_views = inf.get('endorsement_avg_views', 0)
        avg_views_int = inf.get('avg_views_int', 0)
        
        rate_reel = item.get('rate_reel', 0)
        rate_story = item.get('rate_story', 0)
        rate_feed = item.get('rate_feed', 0)
        
        # CPV & CPE calculation
        views_val = endo_avg_views or avg_views_int or 0
        cpv = round(rate_reel / views_val) if (views_val > 0 and rate_reel > 0) else None
        
        # CPE calculation
        endo_likes = inf.get('endorsement_avg_likes', 0)
        endo_comments = inf.get('endorsement_avg_comments', 0)
        
        if not endo_likes:
            raw_likes = inf.get('avg_likes', '0')
            try:
                if 'K' in str(raw_likes):
                    endo_likes = int(float(str(raw_likes).replace('K', '')) * 1000)
                elif 'M' in str(raw_likes):
                    endo_likes = int(float(str(raw_likes).replace('M', '')) * 1000000)
                else:
                    endo_likes = int(float(raw_likes))
            except ValueError:
                endo_likes = 0
                
        if not endo_comments:
            try:
                endo_comments = int(inf.get('avg_comments', 0))
            except ValueError:
                endo_comments = 0
                
        eng_val = endo_likes + endo_comments
        cpe = round(rate_reel / eng_val) if (eng_val > 0 and rate_reel > 0) else None
        
        row_data = [
            f"@{username}",
            inf.get('name', ''),
            followers_int if followers_int else None,
            er_val,
            endo_posts if endo_posts else None,
            endo_er_val,
            endo_avg_views if endo_avg_views else None,
            avg_views_int if avg_views_int else None,
            rate_reel if rate_reel else None,
            rate_story if rate_story else None,
            rate_feed if rate_feed else None,
            cpv,
            cpe,
            ', '.join(inf.get('tags', [])) if inf.get('tags') else None,
            None if is_client else (inf.get('email', '') if inf.get('email') else None),
            None if is_client else (inf.get('phone', '') if inf.get('phone') else None),
            inf.get('location', '') if inf.get('location') else None,
            None if is_client else item.get('status', 'Scouted'),
            None if is_client else (item.get('notes', '') if item.get('notes') else None)
        ]
        
        ws.append(row_data)
        ws.row_dimensions[row_idx].height = 20
        
        # Apply style to data cell
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.font = data_font
            cell.border = thin_border
            
            # Format styles & alignments
            header_name = headers[col_num - 1]
            val = row_data[col_num - 1]
            
            if header_name in ['Username', 'Status CRM']:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif header_name in ['Nama', 'Niche', 'Email', 'No WA', 'Lokasi', 'Catatan']:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                
            # Number formatting
            if val is not None:
                if 'Rate' in header_name or 'CPV' in header_name or 'CPE' in header_name:
                    cell.number_format = '"Rp"#,##0'
                elif 'Followers' in header_name or 'Views' in header_name or 'Sponsor Posts' in header_name:
                    cell.number_format = '#,##0'
                elif 'ER' in header_name:
                    cell.number_format = '0.0%'
                    
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 4:  # Skip title row length calculations
                continue
            if cell.value is not None:
                val_str = str(cell.value)
                if len(val_str) > max_len:
                    max_len = len(val_str)
        # Apply width with padding
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Write to a buffer and return
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"KOL_Scout_{folder['name'].replace(' ', '_')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

# ============================================================================
# NEW FEATURE ROUTES: ENDORSEMENT CRM, PROPOSAL BUILDER, CAMPAIGN WORKSPACE, POST TRACKER
# ============================================================================

@app.route('/update_rate_card', methods=['POST'])
def update_rate_card():
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
    data = request.get_json() or {}
    list_id = data.get('list_id')
    username = data.get('username')
    rate_reel = data.get('rate_reel', 0)
    rate_story = data.get('rate_story', 0)
    rate_feed = data.get('rate_feed', 0)
    
    if not list_id or not username:
        return jsonify({'status': 'error', 'message': 'Missing parameters'}), 400
        
    folders = database.get_user_lists(session['user_id'])
    if not any(str(f['id']) == str(list_id) for f in folders):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
        
    success = database.update_item_rates(list_id, username, rate_reel, rate_story, rate_feed)
    if success:
        return jsonify({'status': 'success'})
    return jsonify({'status': 'error', 'message': 'Failed to update rates'}), 500


@app.route('/campaign/<int:list_id>')
def campaign_workspace(list_id):
    if 'user_id' not in session:
        return redirect(url_for('login', next=request.path))
        
    user_id = session['user_id']
    folders = database.get_user_lists(user_id)
    folder = next((f for f in folders if f['id'] == list_id), None)
    if not folder:
        return "Access denied", 403
        
    items = database.get_list_items(list_id)
    all_influencers = load_data()
    influencer_map = {i['username']: i for i in all_influencers}
    
    merged_items = []
    for item in items:
        username = item['username']
        if username in influencer_map:
            infl_detail = dict(influencer_map[username])
            infl_detail.update(item)
            merged_items.append(infl_detail)
            
    return render_template('campaign_workspace.html',
                           folders=folders,
                           active_folder=folder,
                           influencers=merged_items)


@app.route('/campaign/update_field', methods=['POST'])
def campaign_update_field():
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
    data = request.get_json() or {}
    list_id = data.get('list_id')
    username = data.get('username')
    field_name = data.get('field_name')
    value = data.get('value')
    
    if not list_id or not username or not field_name:
        return jsonify({'status': 'error', 'message': 'Missing parameters'}), 400
        
    folders = database.get_user_lists(session['user_id'])
    if not any(str(f['id']) == str(list_id) for f in folders):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
        
    success = database.update_item_campaign_field(list_id, username, field_name, value)
    if success:
        return jsonify({'status': 'success'})
    return jsonify({'status': 'error', 'message': 'Failed to update field'}), 500


@app.route('/proposal/<int:list_id>')
def proposal_view(list_id):
    if 'user_id' not in session:
        return redirect(url_for('login', next=request.path))
    user_id = session['user_id']
    folders = database.get_user_lists(user_id)
    folder = next((f for f in folders if f['id'] == list_id), None)
    if not folder:
        return "Access denied", 403
    return redirect(url_for('proposal_share', list_id=list_id))


@app.route('/proposal/share/<int:list_id>')
def proposal_share(list_id):
    with database.get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM saved_lists WHERE id = ?", (list_id,))
        folder = cursor.fetchone()
        
    if not folder:
        return "Proposal not found", 404
        
    folder = dict(folder)
    items = database.get_list_items(list_id)
    all_influencers = load_data()
    influencer_map = {i['username']: i for i in all_influencers}
    
    merged_items = []
    for item in items:
        username = item['username']
        if username in influencer_map:
            infl_detail = dict(influencer_map[username])
            infl_detail['rate_reel'] = item.get('rate_reel', 0)
            infl_detail['rate_story'] = item.get('rate_story', 0)
            infl_detail['rate_feed'] = item.get('rate_feed', 0)
            
            # Hide sensitive fields for client link
            infl_detail['email'] = ''
            infl_detail['phone'] = ''
            infl_detail['notes'] = ''
            infl_detail['status'] = ''
            
            merged_items.append(infl_detail)
            
    return render_template('proposal_view.html',
                           active_folder=folder,
                           influencers=merged_items)


@app.route('/campaign/tracker/<int:list_id>')
def post_tracker(list_id):
    if 'user_id' not in session:
        return redirect(url_for('login', next=request.path))
        
    user_id = session['user_id']
    folders = database.get_user_lists(user_id)
    folder = next((f for f in folders if f['id'] == list_id), None)
    if not folder:
        return "Access denied", 403
        
    items = database.get_list_items(list_id)
    all_influencers = load_data()
    influencer_map = {i['username']: i for i in all_influencers}
    
    kols = []
    for item in items:
        username = item['username']
        if username in influencer_map:
            kols.append({
                'username': username,
                'name': influencer_map[username].get('name', username)
            })
            
    tracked_posts = database.get_tracked_posts(list_id)
    
    return render_template('post_tracker.html',
                           folders=folders,
                           active_folder=folder,
                           kols=kols,
                           tracked_posts=tracked_posts)


@app.route('/campaign/tracker/add', methods=['POST'])
def add_post_to_tracker():
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
    list_id = request.form.get('list_id')
    username = request.form.get('username', '').strip()
    post_url = request.form.get('post_url', '').strip()
    
    if not list_id or not username or not post_url:
        return jsonify({'status': 'error', 'message': 'Missing parameters'}), 400
        
    folders = database.get_user_lists(session['user_id'])
    if not any(str(f['id']) == str(list_id) for f in folders):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
        
    scraper = ApifyInstagramScraper()
    post_data = scraper.scrape_post_by_url(post_url)
    
    if not post_data:
        return jsonify({'status': 'error', 'message': 'Gagal mengambil data postingan. Pastikan post bersifat publik.'}), 500
        
    likes = post_data.get('likes', 0)
    comments = post_data.get('comments', 0)
    views = post_data.get('views', 0)
    caption = post_data.get('caption', '')
    
    all_influencers = load_data()
    infl = next((i for i in all_influencers if i['username'].replace('@','') == username.replace('@','')), None)
    er = "0.0%"
    if infl:
        folls = infl.get('followers_int', 0)
        if folls > 0:
            er_val = ((likes + comments) / folls * 100)
            er = f"{er_val:.2f}%"
            
    success = database.add_tracked_post(list_id, username, post_url, caption, likes, comments, views, er)
    if success:
        return jsonify({'status': 'success'})
    return jsonify({'status': 'error', 'message': 'Gagal menyimpan ke database'}), 500


@app.route('/campaign/tracker/refresh', methods=['POST'])
def refresh_tracker_posts():
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
    list_id = request.form.get('list_id')
    if not list_id:
        return jsonify({'status': 'error', 'message': 'Missing list_id'}), 400
        
    folders = database.get_user_lists(session['user_id'])
    if not any(str(f['id']) == str(list_id) for f in folders):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
        
    tracked = database.get_tracked_posts(list_id)
    scraper = ApifyInstagramScraper()
    all_influencers = load_data()
    influencer_map = {i['username']: i for i in all_influencers}
    
    refreshed_count = 0
    for post in tracked:
        post_url = post['post_url']
        username = post['username']
        post_data = scraper.scrape_post_by_url(post_url)
        if post_data:
            likes = post_data.get('likes', 0)
            comments = post_data.get('comments', 0)
            views = post_data.get('views', 0)
            caption = post_data.get('caption', '')
            
            er = "0.0%"
            if username in influencer_map:
                folls = influencer_map[username].get('followers_int', 0)
                if folls > 0:
                    er_val = ((likes + comments) / folls * 100)
                    er = f"{er_val:.2f}%"
                    
            database.add_tracked_post(list_id, username, post_url, caption, likes, comments, views, er)
            refreshed_count += 1
            
    return jsonify({'status': 'success', 'refreshed': refreshed_count})


@app.route('/campaign/tracker/delete/<int:post_id>', methods=['POST'])
def delete_tracked_post(post_id):
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
        
    with database.get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT list_id FROM tracked_posts WHERE id = ?", (post_id,))
        row = cursor.fetchone()
        
    if not row:
        return jsonify({'status': 'error', 'message': 'Post not found'}), 404
        
    list_id = row['list_id']
    folders = database.get_user_lists(session['user_id'])
    if not any(str(f['id']) == str(list_id) for f in folders):
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
        
    success = database.remove_tracked_post(post_id)
    if success:
        return jsonify({'status': 'success'})
    return jsonify({'status': 'error', 'message': 'Failed to delete'}), 500


@app.route('/influencer/refresh/<username>')
def refresh_influencer(username):
    if 'user_id' not in session:
        return redirect(url_for('login', next=request.path))
        
    try:
        scraper = ApifyInstagramScraper()
        new_data = scraper.scrape_single_profile(username)
        if new_data:
            scraper.save_results([new_data])
            return redirect(request.referrer or url_for('home'))
        else:
            return "Gagal memperbarui profil. Profil tidak ditemukan atau API Limit tercapai.", 500
    except Exception as e:
        return f"Error: {e}", 500


@app.route('/api/search', methods=['POST'])
def api_search():
    return jsonify({"status": "deprecated, use home search"})

# ============================================================================
# ADMIN MODULE
# ============================================================================

@app.route('/admin')
def admin_dashboard():
    if not session.get('is_admin'):
        return redirect(url_for('login'))
    return render_template('admin.html', total_db=len(load_data()))

@app.route('/api/start_scrape', methods=['POST'])
def start_scrape():
    if not session.get('is_admin'): return jsonify({'status': 'unauthorized'}), 401
    import subprocess
    try:
        subprocess.Popen([os.path.join('.', 'venv', 'Scripts', 'python.exe'), 'scrape_malang.py'])
        return jsonify({'status': 'started'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/scraping_status')
def scraping_status():
    if not session.get('is_admin'): return jsonify({'status': 'unauthorized'}), 401
    file_path = 'scraping_status.json'
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return jsonify(json.load(f))
        except: pass
    return jsonify({'status': 'Tidak ada proses berjalan', 'is_running': False})

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
